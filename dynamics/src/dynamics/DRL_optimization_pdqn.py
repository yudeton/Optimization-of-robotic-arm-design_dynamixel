#!/usr/bin/env python3
# coding: utf-8
import importlib
import sys
import yaml
import rospy

importlib.reload(sys)
import argparse
import csv
import math
import time
from collections import namedtuple
from math import pi
from os import path

import geometry_msgs.msg
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import plotly.graph_objs as go
import plotly.offline as py
import roboticstoolbox as rtb
import sympy as sp
# import dyna_space
from interface_control.msg import (cal_cmd, cal_process, cal_result, dyna_data,
                                dyna_space_data, optimal_design,
                                optimal_random, specified_parameter_design, tested_model_name, arm_structure)
from matplotlib import cm
from moveit_msgs.msg import DisplayTrajectory, RobotTrajectory
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter
from plotly.offline import download_plotlyjs, iplot, plot
from scipy.interpolate import make_interp_spline  # draw smooth
from spatialmath import *
from trajectory_msgs.msg import JointTrajectory, JointTrajectoryPoint
# from dqn import DQN

np.set_printoptions(
    linewidth=100,
    formatter={"float": lambda x: f"{x:8.4g}" if abs(x) > 1e-10 else f"{0:8.4g}"},
)

from arm_workspace import arm_workspace_plane
# from robot_urdf import RandomRobot
from motor_module import motor_data
from random_robot import RandomRobot
from modular_robot_6dof import modular_robot_6dof
# DRL_optimization api
import sys
import os
import torch.nn as nn
import torch.nn.functional as F
curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
parent_path = os.path.dirname(curr_path)  # 父路径
sys.path.append(parent_path)  # 添加路径到系统路径

import gym
import torch
import datetime
import numpy as np
from common.utils import save_results, make_dir
from common.utils import plot_rewards
from dqn import DQN

import argparse
import random

import matplotlib.pyplot as plt
# from RobotOptEnv_dynamixel import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
from RobotOptEnv_dynamixel_v2_real_pdqn import RobotOptEnv
import tensorboardX
import yaml

# add
import tensorflow as tf
from tf_agents.agents.dqn.dqn_agent import DqnAgent, DdqnAgent
from tf_agents.networks.q_network import QNetwork
from tf_agents.agents.categorical_dqn import categorical_dqn_agent
from tf_agents.drivers import dynamic_step_driver
from tf_agents.environments import suite_gym
from tf_agents.environments import tf_py_environment
from tf_agents.eval import metric_utils
from tf_agents.metrics import tf_metrics
from tf_agents.networks import categorical_q_network
from tf_agents.policies import random_tf_policy
from tf_agents.replay_buffers import tf_uniform_replay_buffer
from tf_agents.trajectories import trajectory
from tf_agents.utils import common
from tf_agents.policies import policy_saver 
from tf_agents.policies import py_tf_eager_policy 

# PDQNPolicy ++
from tf_agents.policies import TFPolicy # add
from tf_agents.trajectories import policy_step
from tf_agents.trajectories.time_step import TimeStep
from tf_agents.specs import tensor_spec


# pdqn
from tensorflow.keras import layers, Model

# add
import io
import os
import shutil
import tempfile
import zipfile

import itertools

files = None
# add

file_path = curr_path + "/outputs/"
curr_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")  # 获取当前时间
curr_time_real = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")  # 获取当前时间
# curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
# tempdir = curr_path + "/C51_outputs/" + \
#             '/' + curr_time + '/models/'  # 保存模型的路径
# tb = tensorboardX.SummaryWriter()
tb = None
# tensorboard_callback = tensorboardX.(log_dir=log_dir, histogram_freq=1)
algo_name = "C51"  # 算法名称
env_name = 'C51_RobotOptEnv'  # 环境名称

#-----------dpqn-----------

#Q網路和參數網路
class QNet(Model):
    def __init__(self, input_dim, action_dim):
        super(QNet, self).__init__()
        self.fc1 = layers.Dense(256, activation='relu')
        self.fc2 = layers.Dense(128, activation='relu')
        self.fc3 = layers.Dense(64, activation='relu')
        self.out = layers.Dense(action_dim)

    def call(self, x):
        x = self.fc1(x)
        x = self.fc2(x)
        x = self.fc3(x)
        return self.out(x)

class ParamNet(Model):
    def __init__(self, input_dim, param_dim):
        super(ParamNet, self).__init__()
        self.fc1 = layers.Dense(256, activation='relu')
        self.fc2 = layers.Dense(128, activation='relu')
        self.fc3 = layers.Dense(64, activation='relu')
        self.out = layers.Dense(param_dim, activation='tanh')

    def call(self, x):
        x = self.fc1(x)
        x = self.fc2(x)
        x = self.fc3(x)
        return self.out(x)
    
#定义一个经验回放缓冲区来存储状态转移
class ReplayBuffer:
    def __init__(self, max_size, input_shape, n_actions, n_params):
        self.mem_size = max_size
        self.mem_cntr = 0
        self.state_memory = np.zeros((self.mem_size, *input_shape), dtype=np.float32)
        self.new_state_memory = np.zeros((self.mem_size, *input_shape), dtype=np.float32)
        self.action_memory = np.zeros(self.mem_size, dtype=np.int32)
        self.param_memory = np.zeros((self.mem_size, n_params), dtype=np.float32)
        self.reward_memory = np.zeros(self.mem_size, dtype=np.float32)
        self.terminal_memory = np.zeros(self.mem_size, dtype=np.bool)

    def store_transition(self, state, action, params, reward, state_):
        index = self.mem_cntr % self.mem_size
        self.state_memory[index] = state
        self.new_state_memory[index] = state_
        self.action_memory[index] = action
        self.param_memory[index] = params
        self.reward_memory[index] = reward
        self.terminal_memory[index] = 0

        self.mem_cntr += 1

    def sample_buffer(self, batch_size):
        max_mem = min(self.mem_cntr, self.mem_size)
        batch = np.random.choice(max_mem, batch_size, replace=False)

        states = self.state_memory[batch]
        states_ = self.new_state_memory[batch]
        rewards = self.reward_memory[batch]
        actions = self.action_memory[batch]
        params = self.param_memory[batch]
        terminals = self.terminal_memory[batch]

        return states, actions, params, rewards, states_

class PDQNPolicy(TFPolicy):
    def __init__(self, q_net, param_net, time_step_spec, action_spec):
        self._q_net = q_net
        self._param_net = param_net
        super(PDQNPolicy, self).__init__(time_step_spec, action_spec)

    def _action(self, time_step: TimeStep, policy_state, seed):
        state = time_step.observation
        params = self._param_net(state)
        q_input = tf.concat([state, params], axis=1)
        q_values = self._q_net(q_input)
        action = tf.argmax(q_values, axis=1, output_type=tf.int32)
        return policy_step.PolicyStep(action, policy_state)

    def _distribution(self, time_step: TimeStep, policy_state):
        raise NotImplementedError("This policy does not support action distributions.")


#定义 PDQN 代理，包括选择动作和学习过程
class PDQNAgent:
    def __init__(self, state_dim, action_dim, param_dim, buffer_size=100000, batch_size=64, gamma=0.99, lr_q=0.001, lr_p=0.001, log_dir=None, model_path=None):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.param_dim = param_dim
        self.buffer = ReplayBuffer(buffer_size, (state_dim,), action_dim, param_dim)
        self.batch_size = batch_size
        self.gamma = gamma

        self.q_net = QNet(state_dim + param_dim, action_dim)
        self.target_q_net = QNet(state_dim + param_dim, action_dim)
        self.param_net = ParamNet(state_dim, param_dim)

        self.optimizer_q = tf.keras.optimizers.Adam(learning_rate=lr_q) #配置優化器
        self.optimizer_p = tf.keras.optimizers.Adam(learning_rate=lr_p) #配置優化器
        
        self.update_target_network()    #更新目標網路

        # 设置策略保存路径和策略保存器
        if log_dir:
            self.summary_writer = tf.summary.create_file_writer(log_dir)

    def update_target_network(self):
        self.target_q_net.set_weights(self.q_net.get_weights())

    def choose_action(self, state, epsilon):
        if np.random.random() < epsilon:
            action = np.random.choice(self.action_dim)
            params = np.random.uniform(-1, 1, self.param_dim)
        else:
            state = tf.convert_to_tensor([state], dtype=tf.float32)
            params = self.param_net(state)
            q_input = tf.concat([state, params], axis=1)
            q_values = self.q_net(q_input)
            action = tf.argmax(q_values[0]).numpy()
            params = params.numpy()[0]
        return action, params

    # 将轨迹添加到重放缓冲区
    def store_transition(self, state, action, params, reward, state_):
        self.buffer.store_transition(state, action, params, reward, state_)

    def learn(self):
        if self.buffer.mem_cntr < self.batch_size:
            return None, None

        states, actions, params, rewards, states_ = self.buffer.sample_buffer(self.batch_size)
        states = tf.convert_to_tensor(states, dtype=tf.float32)
        states_ = tf.convert_to_tensor(states_, dtype=tf.float32)
        rewards = tf.convert_to_tensor(rewards, dtype=tf.float32)

        with tf.GradientTape() as tape:
            params_ = self.param_net(states_)
            q_next = self.target_q_net(tf.concat([states_, params_], axis=1))
            q_next = tf.reduce_max(q_next, axis=1)
            q_target = rewards + self.gamma * q_next
            q_target = tf.reshape(q_target, (self.batch_size, 1))

            actions_onehot = tf.one_hot(actions, self.action_dim)
            q_eval = self.q_net(tf.concat([states, params], axis=1))
            q_eval = tf.reduce_sum(q_eval * actions_onehot, axis=1, keepdims=True)
            
            loss_q = tf.reduce_mean(tf.keras.losses.MSE(q_target, q_eval))
            #loss_q = tf.keras.losses.MSE(q_target, q_eval)

        grads = tape.gradient(loss_q, self.q_net.trainable_variables)
        self.optimizer_q.apply_gradients(zip(grads, self.q_net.trainable_variables))

        with tf.GradientTape() as tape:
            x_all = self.param_net(states)
            q_all = self.q_net(tf.concat([states, x_all], axis=1))
            loss_p = -tf.reduce_mean(tf.reduce_sum(q_all, axis=1))

        grads = tape.gradient(loss_p, self.param_net.trainable_variables)
        self.optimizer_p.apply_gradients(zip(grads, self.param_net.trainable_variables))

        self.update_target_network()

        return loss_q.numpy(), loss_p.numpy()

    def save_model(self, episode):
        
        '''
        filename = f'policy_step{episode}'
        self.train_checkpointer.save(global_step=episode)  # 保存检查点
        self.tf_policy_saver.save(self.policy_dir + '/' + filename)  # 保存策略
        '''
        model_dir = f"models/episode_{episode}"
        if not os.path.exists(model_dir):
            os.makedirs(model_dir)
        self.q_net.save_weights(os.path.join(model_dir, 'q_net.h5'))
        self.param_net.save_weights(os.path.join(model_dir, 'param_net.h5'))
        


#訓練過程

def train_pdqn(agent, env, model_path, episodes=20000, epsilon_start=1.0, epsilon_end=0.1, epsilon_decay=0.995, save_interval=10000):
    epsilon = epsilon_start
    for episode in range(episodes):     #訓練的總迭代次數
        state = env.reset()
        done = False
        total_reward = 0
        step_count = 0

        while not done:
            action, params = agent.choose_action(state, epsilon)
            next_state, reward, done, _ = env.step((action, params))
            agent.store_transition(state, action, params, reward, next_state)   # 在当前环境状态下使用指定策略执行一步操作，并将收集的轨迹添加到重放缓冲区
            loss_q, loss_p = agent.learn()
            state = next_state
            total_reward += reward
            step_count += 1

        epsilon = max(epsilon_end, epsilon * epsilon_decay)

        print(f'Episode: {episode}, Total Reward: {total_reward}, Steps: {step_count}, Epsilon: {epsilon:.3f}, Q Loss: {loss_q}, Param Loss: {loss_p}')

        with agent.summary_writer.as_default():
            tf.summary.scalar('Episode_Return', total_reward, step=episode)
            if loss_q is not None and loss_p is not None:
                tf.summary.scalar('Loss/Q_loss', loss_q, step=episode)
                tf.summary.scalar('Loss/Param_loss', loss_p, step=episode)
            agent.summary_writer.flush()

        if episode % save_interval == 0:
            agent.save_model(episode)   # 设置模型保存路径和检查点


#-----------dpqn-----------

class drl_optimization:
    def __init__(self):
        # 初始化機器人和環境
        self.robot = modular_robot_6dof()
        self.env = RobotOptEnv()

    def env_agent_config(self, cfg, algorithm, seed=1):
        ''' 创建环境和智能体'''

        # 设置神经网络的层参数
        fc_layer_params = (100,)
        learning_rate = 1e-3    # 學習率
        gamma = 0.99            # 折扣因子
        num_atoms = 51          # @param {type:"integer"}
        min_q_value = -100      # @param {type:"integer"}
        max_q_value = 50        # @param {type:"integer"}
        n_step_update = 2       # N步更新

        # 包裝環境
            # 將gym環境包成TensorFlow兼容的環境
        train_py_env = suite_gym.wrap_env(self.env)
        env = tf_py_environment.TFPyEnvironment(train_py_env)

        categorical_q_net = categorical_q_network.CategoricalQNetwork(
            env.observation_spec(),
            env.action_spec(),
            num_atoms=num_atoms,
            fc_layer_params=fc_layer_params)

        dqn_network = QNetwork(
            env.observation_spec(),
            env.action_spec(),
            fc_layer_params=fc_layer_params)

        ddqn_network = QNetwork(
            env.observation_spec(),
            env.action_spec(),
            fc_layer_params=fc_layer_params)

        # 配置優化器
        optimizer = tf.compat.v1.train.AdamOptimizer(learning_rate=learning_rate)
        self.global_step = tf.compat.v1.train.get_or_create_global_step()

        # 根据指定的算法初始化智能体
        if algorithm == 'DQN':
            agent = DqnAgent(
                env.time_step_spec(),
                env.action_spec(),
                q_network = dqn_network,
                optimizer = optimizer,
                n_step_update=n_step_update,
                td_errors_loss_fn = common.element_wise_squared_loss,
                gamma=gamma,
                train_step_counter = self.global_step)
            agent.initialize()
            rospy.loginfo("DRL algorithm init: %s", algorithm)
        elif algorithm == 'DDQN':
            agent = DdqnAgent(
                env.time_step_spec(),
                env.action_spec(),
                q_network = ddqn_network,
                optimizer = optimizer,
                n_step_update=n_step_update,
                td_errors_loss_fn = common.element_wise_squared_loss,
                gamma=gamma,
                train_step_counter = self.global_step)
            agent.initialize()
            rospy.loginfo("DRL algorithm init: %s", algorithm)
        elif algorithm == 'C51':
            agent = categorical_dqn_agent.CategoricalDqnAgent(
                env.time_step_spec(),
                env.action_spec(),
                categorical_q_network=categorical_q_net,
                optimizer=optimizer,
                min_q_value=min_q_value,
                max_q_value=max_q_value,
                n_step_update=n_step_update,
                td_errors_loss_fn=common.element_wise_squared_loss,
                gamma=gamma,
                train_step_counter=self.global_step)
            agent.initialize()
            rospy.loginfo("DRL algorithm init: %s", algorithm)

        # 返回環境和智能體
        return env, agent

class RosTopic:
    def __init__(self):
        self.sub_taskcmd = rospy.Subscriber("/cal_command", cal_cmd, self.cmd_callback)
        self.sub_test_model_name = rospy.Subscriber("/tested_model_name", tested_model_name, self.test_model_name_callback)
        # Subscriber select dof and structure
        self.sub_arm_structure = rospy.Subscriber("/arn_structure", arm_structure, self.arm_structure_callback)
        self.cmd_run = 0
        self.arm_structure_dof = None
        self.DRL_algorithm = None
    def cmd_callback(self, data):
        self.cmd = data.cmd
        rospy.loginfo("I heard command is %s", data.cmd)
        if data.cmd == 22:
            rospy.sleep(10)
            self.cmd_run = 1
        if data.cmd == 23:
            rospy.sleep(10)
            self.cmd_run = 2
    def test_model_name_callback(self, data):
        self.test_model_name  = data.tested_model_name
        print(self.test_model_name)

    def arm_structure_callback(self, data):
        self.arm_structure_dof = data.DoF
        self.DRL_algorithm = data.structure_name
        # rospy.loginfo("DoF: %s", self.arm_structure_dof)
        # rospy.loginfo("DRL algorithm: %s", self.DRL_algorithm)

class Trainer:
    def __init__(self, agent, env, model_path):
        self.agent = agent
        self.env = env
        self.model_path = model_path
        # TODO: fixed

        self.num_iterations = 60000 # 训练的总迭代次数
        self.initial_collect_steps = 1000  # 初始数据收集的步数
        self.collect_steps_per_iteration = 1  # 每次迭代的数据收集步数
        self.replay_buffer_capacity = 100000  # 重放缓冲区的容量

         # 初始化随机策略，用于初始数据收集
        self.random_policy = random_tf_policy.RandomTFPolicy(self.env.time_step_spec(),
                                                self.env.action_spec())
        
        # 初始化重放缓冲区，用于存储采集的数据
        self.replay_buffer = tf_uniform_replay_buffer.TFUniformReplayBuffer(
            data_spec=self.agent.collect_data_spec,
            batch_size=self.env.batch_size,
            max_length=self.replay_buffer_capacity)
        
        # 配置数据收集驱动，每次收集指定步数的数据
        self.collect_driver = dynamic_step_driver.DynamicStepDriver(
            self.env,
            self.agent.collect_policy,
            observers=[self.replay_buffer.add_batch],
            num_steps=self.collect_steps_per_iteration)
        
        # 运行初始数据收集
        self.collect_driver.run()

        # 设置模型保存路径和检查点
        self.checkpoint_dir = os.path.join(self.model_path, 'checkpoint')
        self.train_checkpointer = common.Checkpointer(
            ckpt_dir=self.checkpoint_dir,
            max_to_keep=1,
            agent=self.agent,
            policy=self.agent.policy,
            replay_buffer=self.replay_buffer,
            global_step=self.agent.train_step_counter
        )

        # 初始化或恢复检查点
        self.train_checkpointer.initialize_or_restore()

        # 设置策略保存路径和策略保存器
        self.policy_dir = os.path.join(self.model_path, 'policy')
        self.tf_policy_saver = policy_saver.PolicySaver(agent.policy)
        
        # 设定其他训练参数
        self.batch_size = 64                # 每个批次的数据量
        self.n_step_update = 2              # N步更新
        self.num_eval_episodes = 10         # 每次评估的回合数
        self.log_interval = 200             # 日志记录的间隔
        self.eval_interval = 1000           # 评估的间隔
        self.checkpoint_interval = 10000    # 检查点保存的间隔

    def collect_step(self, environment, policy):
        # 在当前环境状态下使用指定策略执行一步操作，并将收集的轨迹添加到重放缓冲区
        time_step = environment.current_time_step()
        action_step = policy.action(time_step)
        next_time_step = environment.step(action_step.action)
        traj = trajectory.from_transition(time_step, action_step, next_time_step)

        # 将轨迹添加到重放缓冲区
        self.replay_buffer.add_batch(traj)

        return next_time_step

    def train(self, pre_fr=0, train_eps = 300):
        rospy.loginfo("-------------數據採集------------")

        # 进行初始数据收集
        for _ in range(self.initial_collect_steps):
            
            time_step_collect = self.collect_step(self.env, self.random_policy)
            rospy.loginfo("initial_collect_steps: %s", _)
            collect_step_reward = time_step_collect.reward.numpy()[0]
            collect_step_state = time_step_collect.observation.numpy()[0]
            rospy.loginfo("collect_step_reward: %s", collect_step_reward)
            rospy.loginfo("collect_step_state: %s", collect_step_state)

            if time_step_collect.is_last():
                time_step_collect = self.env.reset()

        # 设置数据集以供训练使用
        dataset = self.replay_buffer.as_dataset(
            num_parallel_calls=3, sample_batch_size=self.batch_size,
            num_steps=self.n_step_update + 1).prefetch(3)

        iterator = iter(dataset)

        # 将智能体的训练函数包装在TF函数中，以优化性能
        self.agent.train = common.function(self.agent.train)

        # 重置训练步骤计数器
        self.agent.train_step_counter.assign(0)
        rospy.loginfo("Evaluate the agent's policy once before training.")

        rospy.loginfo("-------------Train Start------------")
        episode_return = 0.0
        
        # 進行訓練迭代 60000
        for _ in range(self.num_iterations):
        # 每次迭代中收集指定步数的数据 1
            for _ in range(self.collect_steps_per_iteration):
                time_step = self.collect_step(self.env, self.agent.collect_policy)
            
            # 从重放缓冲区采样批次数据进行训练
            experience, unused_info = next(iterator)
            train_loss = self.agent.train(experience)
            step = self.agent.train_step_counter.numpy()
            tb.add_scalar("/trained-model/Loss_per_frame/", float(train_loss.loss), step)
            
            # 開始tensorboard紀錄，记录和打印训练过程中的相关信息
            step_reward = time_step.reward.numpy()[0]
            step_state = time_step.observation.numpy()[0]
            tb.add_scalar("/trained-model/train_step_reward/", step_reward, step)
            episode_return += step_reward
            rospy.loginfo("step_reward: %s", step_reward)
            rospy.loginfo("step_state: %s", step_state)
            rospy.loginfo("step: %s", step)
            rospy.loginfo("train_loss: %s", float(train_loss.loss))
            rospy.loginfo("================================")

            if time_step.is_last():
                time_step = self.env.reset()
                rospy.loginfo("episode_return: %s", episode_return)
                rospy.loginfo("------episode end.-----------")
                tb.add_scalar("/trained-model/Episode_Return/", episode_return, step)
                episode_return = 0.0

            # 定期打印日志信息
            if step % self.log_interval == 0:
                print('step = {0}: loss = {1}'.format(step, train_loss.loss))
                tb.add_scalar("/trained-model/loss_log/",  float(train_loss.loss), step)

            # 定期保存检查点和模型
            if step % self.checkpoint_interval == 0:
                filename = 'policy_step{}'.format(self.agent.train_step_counter.numpy())
                self.train_checkpointer.save(self.agent.train_step_counter)
                self.tf_policy_saver.save(self.policy_dir+ '/' + filename)

        # 保存最終的訓練模型
        self.train_checkpointer.save(self.agent.train_step_counter)
        self.train_checkpointer.initialize_or_restore()
        self.agent.global_step = tf.compat.v1.train.get_global_step()
        self.tf_policy_saver.save(self.policy_dir)


class Tester(object):

    def __init__(self, env, model_path, drl_env_class, num_episodes=50, max_ep_steps=400, test_ep_steps=100):
        self.num_episodes = num_episodes
        self.max_ep_steps = max_ep_steps
        self.test_ep_steps = test_ep_steps
        # self.agent = agent
        self.model_path = model_path
        self.env = env
        self._best_episode_reward = 50
        self.drl_env_class = drl_env_class
        # self.agent.is_training = False
        # self.agent.load_weights(model_path)
        # self.policy = lambda x: agent.act(x)
        self.xlsx_outpath = "./xlsx/"

    def test(self, debug=True, visualize=True): # debug = true
        print("load model ")
        self.policy_dir = os.path.join(self.model_path, 'policy')
        saved_policy = tf.saved_model.load(self.policy_dir)

        excel_file = Workbook()
        sheet = excel_file.active
        i = 0
        # avg_reward = 0
        step = 0
        episode = 0
        avg_reward = 0
        state_traj = []
        motor_config_state_traj = []
        num = 0
        for _ in range(self.num_episodes):
            time_step = self.env.reset()
            while not time_step.is_last():
                # step = self.agent.train_step_counter.numpy()
                action_step = saved_policy.action(time_step)
                time_step = self.env.step(action_step.action)
                step_reward = time_step.reward.numpy()[0]
                state = time_step.observation
                state_traj.append(state.numpy()[0])
                motor_config_state_traj.append([self.drl_env_class.motor_type_axis_2, self.drl_env_class.motor_type_axis_3])
                step += 1
                tb.add_scalar("/tested-model/test_step_reward/", step_reward, step)
                rospy.loginfo("step_reward: %s", step_reward)
                rospy.loginfo("step_state: %s", state.numpy()[0])
                rospy.loginfo("step: %s", step)
            
            episode += 1
            episode_reward = step_reward
            rospy.loginfo("episode_reward: %s", episode_reward)
            rospy.loginfo("================================")
            for j in range(len(state.numpy()[0])):
                sheet.cell(row=i+1, column=j+1).value = state.numpy()[0][j] # state record 
            # self.motor_rated[1], self.motor_rated[2]] record
            sheet.cell(row=i+1, column=len(state.numpy()[0])+4).value = self.drl_env_class.motor_type_axis_2
            sheet.cell(row=i+1, column=len(state.numpy()[0])+5).value = self.drl_env_class.motor_type_axis_3
            
            # reward record
            sheet.cell(row=i + 1, column=len(state.numpy()[0])+2).value = episode_reward
            i = i + 1

            tb.add_scalar("/tested-model/test_episode_reward/", episode_reward, episode)
            avg_reward += episode_reward
            # TODO: 若獎勵大於200, 則儲存當前回合獎勵軌跡 
            if episode_reward >= self._best_episode_reward: 
                # TODO: state_traj
                excel_file_traj = Workbook()
                sheet_traj = excel_file_traj.active
                # 迭代矩陣的每一個元素，並填入工作表中
                for k in range(len(state_traj)):
                    for l in range(len(state_traj[k])):
                        sheet_traj.cell(row=k+1, column=l+1).value = state_traj[k][l]
                for m in range(len(motor_config_state_traj)):
                    for n in range(len(motor_config_state_traj[m])):
                        sheet_traj.cell(row=m+1, column=n+9).value = motor_config_state_traj[m][n]
                num = num + 1  # 要更改的数字
                new_str = "_{}_".format(num)
                if not os.path.exists(self.model_path + "tested_state_traj_" + curr_time_real +"/"):
                    os.makedirs(self.model_path + "tested_state_traj_" + curr_time_real +"/")
                file_name_traj = self.model_path + "tested_state_traj_" + curr_time_real +"/tested_state_traj" + new_str + curr_time +".xlsx"
                excel_file_traj.save(file_name_traj)
            # TODO: 清空
            else:
                state_traj = []

        avg_reward /= self.num_episodes
        # print("avg reward: %5f" % (avg_reward))
        rospy.loginfo('avg reward: {}'.format(avg_reward))

        # model_path = curr_path + '/train_results' + '/C51_outputs/' + str(arm_structure_dof) + \
        #         '/' + str(ros_topic.test_model_name) + '/models/'  # 保存模型的路径
        
        # add curr_time
        if not os.path.exists(self.model_path + "tested_reward_state_" + curr_time_real +"/"):
                    os.makedirs(self.model_path + "tested_reward_state_" + curr_time_real +"/")
        file_name = self.model_path + "tested_reward_state_" + curr_time_real +"/tested_reward_state_" + curr_time +".xlsx"
        excel_file.save(file_name)

        
        # TODO: save robot model structure, DH, urdf, stl
class ReTrainer:
    def __init__(self, agent, env, model_path, iterations, retrain_path=None):
        self.agent = agent
        self.env = env
        self.model_path = model_path
        self.retrain_path = retrain_path
        # TODO: fixed
        # self.num_iterations = 15000 # @param {type:"integer"}
        # test
        self.num_iterations = iterations # @param {type:"integer"}
        # self.initial_collect_steps = 1000  # @param {type:"integer"}
        # test
        self.initial_collect_steps = 1000  # @param {type:"integer"}
        self.collect_steps_per_iteration = 1  # @param {type:"integer"}
        self.replay_buffer_capacity = 100000  # @param {type:"integer"}

        self.random_policy = random_tf_policy.RandomTFPolicy(self.env.time_step_spec(),
                                                self.env.action_spec())

        self.replay_buffer = tf_uniform_replay_buffer.TFUniformReplayBuffer(
            data_spec=self.agent.collect_data_spec,
            batch_size=self.env.batch_size,
            max_length=self.replay_buffer_capacity)

        self.collect_driver = dynamic_step_driver.DynamicStepDriver(
            self.env,
            self.agent.collect_policy,
            observers=[self.replay_buffer.add_batch],
            num_steps=self.collect_steps_per_iteration)
        # Initial data collection
        self.collect_driver.run()

        self.checkpoint_dir = os.path.join(self.model_path, 'checkpoint')
        self.train_checkpointer = common.Checkpointer(
            ckpt_dir=self.checkpoint_dir,
            max_to_keep=1,
            agent=self.agent,
            policy=self.agent.policy,
            replay_buffer=self.replay_buffer,
            global_step=self.agent.train_step_counter
        )
        self.train_checkpointer.initialize_or_restore()
        self.policy_dir = os.path.join(self.model_path, 'policy')
        self.tf_policy_saver = policy_saver.PolicySaver(self.agent.policy)
        # add -----
        self.batch_size = 64  # @param {type:"integer"}
        self.n_step_update = 2  # @param {type:"integer"}
        self.num_eval_episodes = 10  # @param {type:"integer"}
        self.log_interval = 200  # @param {type:"integer"}
        self.eval_interval = 1000  # @param {type:"integer"}
        self.checkpoint_interval = 10000
        # FIXME: retrain init -----
        # Load the saved policy
        if retrain_path != None:
            saved_policy_path = os.path.join(self.retrain_path, 'policy')
            saved_policy = tf.saved_model.load(saved_policy_path)
            self.tf_policy_saver = saved_policy
            # TODO:
        else:
            print("The model to be retrained is not loaded.")
        # FIXME: TFUniformReplayBuffer 為空，緩衝區中沒有要採樣的項目。 緩衝區需要先充滿經驗，然後才能用於訓練。
        # TODO: use Checkpointer

    def compute_avg_return(self, environment, policy, num_episodes=10):

        total_return = 0.0
        for _ in range(num_episodes):

            time_step = environment.reset()
            episode_return = 0.0

            while not time_step.is_last():
                action_step = policy.action(time_step)
                time_step = environment.step(action_step.action)
                episode_return += time_step.reward
            total_return += episode_return

        avg_return = total_return / num_episodes
        return avg_return.numpy()[0]

    def collect_step(self, environment, policy):
        time_step = environment.current_time_step()
        action_step = policy.action(time_step)
        next_time_step = environment.step(action_step.action)
        traj = trajectory.from_transition(time_step, action_step, next_time_step)

        # Add trajectory to the replay buffer
        self.replay_buffer.add_batch(traj)

        return next_time_step

    def train(self, pre_fr=0, train_eps = 300):
        rospy.loginfo("-------------數據採集------------")
        for _ in range(self.initial_collect_steps):
            time_step_collect = self.collect_step(self.env, self.random_policy)
            rospy.loginfo("initial_collect_steps: %s", _)
            # This loop is so common in RL, that we provide standard implementations of
            # these. For more details see the drivers module.

            # Dataset generates trajectories with shape [BxTx...] where
            # T = n_step_update + 1.
        
        dataset = self.replay_buffer.as_dataset(
            num_parallel_calls=3, sample_batch_size=self.batch_size,
            num_steps=self.n_step_update + 1).prefetch(3)

        iterator = iter(dataset)
        # (Optional) Optimize by wrapping some of the code in a graph using TF function.
        self.agent.train = common.function(self.agent.train)

        # Reset the train step
        self.agent.train_step_counter.assign(0)
        rospy.loginfo("Evaluate the agent's policy once before training.")
        # Evaluate the agent's policy once before training.
        # avg_return = self.compute_avg_return(self.env, self.agent.policy, self.num_eval_episodes)
        # returns = [avg_return]
        rospy.loginfo("-------------Retrain Start------------")
        for _ in range(self.num_iterations):

        # Collect a few steps using collect_policy and save to the replay buffer.
            for _ in range(self.collect_steps_per_iteration):
                time_step = self.collect_step(self.env, self.agent.collect_policy)

            # Sample a batch of data from the buffer and update the agent's network.
            experience, unused_info = next(iterator)
            train_loss = self.agent.train(experience)
            step = self.agent.train_step_counter.numpy()
            tb.add_scalar("/trained-model/Loss_per_frame/", float(train_loss.loss), step)
            # 開始tensorboard紀錄
            step_reward = time_step.reward.numpy()[0]
            step_state = time_step.observation.numpy()[0]
            
            tb.add_scalar("/trained-model/train_step_reward/", step_reward, step)
            # tb.add_scalar("/trained-model/Average_Return/", avg_return, step)
            episode_return += step_reward
            rospy.loginfo("step_reward: %s", step_reward)
            rospy.loginfo("step_state: %s", step_state)
            rospy.loginfo("step: %s", step)
            rospy.loginfo("train_loss: %s", float(train_loss.loss))
            rospy.loginfo("================================")
            if time_step.is_last():
                rospy.loginfo("episode_return: %s", episode_return)
                rospy.loginfo("------episode end.-----------")
                tb.add_scalar("/trained-model/Episode_Return/", episode_return, step)
                episode_return = 0.0

            if step % self.log_interval == 0:
                print('step = {0}: loss = {1}'.format(step, train_loss.loss))
                tb.add_scalar("/trained-model/loss_log/",  float(train_loss.loss), step)
            # if step % self.eval_interval == 0:
            #     avg_return = self.compute_avg_return(self.env, self.agent.policy, self.num_eval_episodes)
            #     print('step = {0}: Average Return = {1:.2f}'.format(step, avg_return))
            #     tb.add_scalar("/trained-model/Average_Return/", avg_return, step)
            #     # returns.append(avg_return)
            # 每一萬步儲存一次 model 
            # self.checkpoint_interval = 2
            if step % self.checkpoint_interval == 0:
                filename = 'policy_step{}'.format(self.agent.train_step_counter.numpy())
                self.train_checkpointer.save(self.agent.train_step_counter)
                self.tf_policy_saver.save(self.policy_dir+ '/' + filename)

        # end save final train model 
        # filename = 'policy_step{}'.format(self.agent.train_step_counter.numpy())
        self.train_checkpointer.save(self.agent.train_step_counter)
        self.train_checkpointer.initialize_or_restore()
        self.agent.global_step = tf.compat.v1.train.get_global_step()

        self.tf_policy_saver.save(self.policy_dir)

if __name__ == "__main__":
    rospy.init_node("optimization")
    a = 0
    # cfg = DQNConfig()
    cfg = 0
    arm_structure_dof = 0
    model_path = None
    drl = drl_optimization()
    # plot_cfg = PlotConfig()
    ros_topic = RosTopic()
    ddqn_train_eps = 1000  # 训练的回合数
    ddqn_test_eps = 100  # 测试的回合数
    
    
    with open(curr_path+'/drl_param.yaml', 'r') as f:
        config = yaml.load(f, Loader=yaml.Loader)
    op_function_flag = config['op_function']
    rospy.loginfo("Input op_function_flag: %s" % op_function_flag)    
    ros_topic.arm_structure_dof = config['arm_structure_dof']
    rospy.loginfo("Input arm_structure_dof: %d" % ros_topic.arm_structure_dof)
    ros_topic.DRL_algorithm = config['DRL_algorithm']
    rospy.loginfo("Input DRL_algorithm: %s" % ros_topic.DRL_algorithm)
    ros_topic.cmd_run = config['cmd_run']
    rospy.loginfo("Input cmd_run: %d" % ros_topic.cmd_run)
    ros_topic.test_model_name = config['test_model_name']
    rospy.loginfo("Input test_model_name: %s" % ros_topic.test_model_name)

    
    # train = Trainer()
    while not rospy.is_shutdown():
        # test all
        if op_function_flag == "case1":
            from RobotOptEnv_dynamixel import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case2":
            from RobotOptEnv_dynamixel_v2 import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case3":
            from RobotOptEnv_dynamixel_v3_motion import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case1_real":
            op_function_flag = "case1"
            from RobotOptEnv_dynamixel_real import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case2_real":
            op_function_flag = "case2"
            from RobotOptEnv_dynamixel_v2_real import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case2_real_pdqn":
            op_function_flag = "case2"
            from RobotOptEnv_dynamixel_v2_real_pdqn import RobotOptEnv
        if ros_topic.arm_structure_dof == 6:
            drl.env = RobotOptEnv()
            rospy.loginfo('arm_structure_dof: {}'.format(ros_topic.arm_structure_dof))
            arm_structure_dof = ros_topic.arm_structure_dof
            ros_topic.arm_structure_dof = 0
        elif ros_topic.arm_structure_dof == 3:
            drl.env = RobotOptEnv_3dof()
            rospy.loginfo('arm_structure_dof: {}'.format(ros_topic.arm_structure_dof))
            arm_structure_dof = ros_topic.arm_structure_dof
            ros_topic.arm_structure_dof = 0
        elif ros_topic.arm_structure_dof == 5:
            drl.env = RobotOptEnv_5dof()
            rospy.loginfo('arm_structure_dof: {}'.format(ros_topic.arm_structure_dof))
            arm_structure_dof = ros_topic.arm_structure_dof
            ros_topic.arm_structure_dof = 0


        drl.env.op_dof = config['op_dof']
        rospy.loginfo("Input op_dof: %d" % drl.env.op_dof)
        
        drl.env.op_payload = config['op_payload']
        # rospy.loginfo("Input op_payload: %d" % drl.env.op_payload)
        rospy.loginfo('Input op_payload_position: {}'.format(drl.env.op_payload))
        
        drl.env.op_payload_position = config['op_payload_position']
        # rospy.loginfo("Input op_payload_position: %d" % drl.env.op_payload_position)
        rospy.loginfo('Input op_payload_position: {}'.format(drl.env.op_payload_position))

        drl.env.mission_time = config['mission_time']
        # rospy.loginfo("Input op_vel: %d" % drl.env.op_vel)
        rospy.loginfo('Input mission_time: {}'.format(drl.env.mission_time))
        
        drl.env.op_vel = config['op_vel']
        # rospy.loginfo("Input op_vel: %d" % drl.env.op_vel)
        rospy.loginfo('Input op_vel: {}'.format(drl.env.op_vel))
        
        drl.env.op_acc = config['op_acc']
        # rospy.loginfo("Input op_acc: %d" % drl.env.op_acc)
        rospy.loginfo('Input op_acc: {}'.format(drl.env.op_acc))
        
        # 輸入可達半徑, 即最大總臂長設定
        drl.env.op_radius = config['op_radius']
        rospy.loginfo("Input op_radius: %d" % drl.env.op_radius)
        
        drl.env.op_weight = config['op_weight']
        rospy.loginfo("Input op_weight: %d" % drl.env.op_weight)
        
        drl.env.op_cost = config['op_cost']
        rospy.loginfo("Input op_cost: %d" % drl.env.op_cost)

        drl.env.action_select = config['action_select']
        rospy.loginfo("Input action_select: %s" % drl.env.action_select)

        drl.env.point_test_excel = config['point_test_excel']
        rospy.loginfo("Input point_test_excel: %s" % drl.env.point_test_excel)

        test_episodes = config['test_episodes']
        rospy.loginfo("Input test_episodes: %d" % test_episodes)

        policy_num = config['policy_num']
        rospy.loginfo("Input policy_num: %d" % policy_num)

        drl.env.train_flag = config['train_flag']
        rospy.loginfo("Input train_flag: %d" % drl.env.train_flag)
        # 要開始時, 按下隨意鍵
        input_text = input("Enter some next: ")
        rospy.loginfo("Input text: %s" % input_text)
        
        if ros_topic.cmd_run == 10:
            tb = tensorboardX.SummaryWriter()
            ros_topic.cmd_run = 0
            if ros_topic.DRL_algorithm == 'DQN':
                model_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' +str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                model_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                model_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'PDQN':
                model_path = curr_path + '/train_results' + '/PDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径

            # 訓練
            drl.env.model_select = "train"
            drl.env.point_Workspace_cal_Monte_Carlo()
            # train_env, train_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=1)
            # # model_path = None
            # train = Trainer(train_agent, train_env, model_path)
            # train.train(train_eps = ddqn_train_eps)
            # # # 測試
            env = RobotOptEnv()
            log_dir = os.path.join("logs", datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
            agent = PDQNAgent(env.state_dim, env.action_dim, env.param_dim, log_dir=log_dir, model_path=model_path)

            # 训练 PDQN 代理
            train_pdqn(agent, env, model_path)
            break


        if ros_topic.cmd_run == 1:
            tb = tensorboardX.SummaryWriter()
            ros_topic.cmd_run = 0
            if ros_topic.DRL_algorithm == 'DQN':
                model_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' +str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                model_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                model_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            
            # 訓練
            drl.env.model_select = "train"
            drl.env.point_Workspace_cal_Monte_Carlo()
            train_env, train_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=1)
            # model_path = None
            train = Trainer(train_agent, train_env, model_path)
            train.train(train_eps = ddqn_train_eps)
            # # # 測試
            # drl.env.model_select = "test"
            # # save_result_path = curr_path + '/test_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
            # #     '/' + curr_time   # 保存結果的路径
            # # plot_cfg.model_path = plot_cfg.model_path +'model_last.pkl'
            # # model_path = model_path + 'policy_step' + str(policy_num) +  '/' # 選擇模型的路径
            # test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
            # test = Tester(test_env, model_path, drl.env, num_episodes = 200) # 20230309  change 300-> 200
            # test.test()
            break

        # test tested_model
        if ros_topic.cmd_run == 2:
            tb = tensorboardX.SummaryWriter()
            ros_topic.cmd_run = 0
            if ros_topic.DRL_algorithm == 'DQN':
                select_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'   # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                select_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                select_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            
            drl.env.model_select = "test"
            model_path = select_path
            test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
            test = Tester(test_env, model_path, drl.env, num_episodes = test_episodes) # 20230309  change 300-> 200 #20230326 mini-test 20
            test.test()
            break
        # retrain 
        if ros_topic.cmd_run == 3:
            tb = tensorboardX.SummaryWriter()
            ros_topic.cmd_run = 0
            if ros_topic.DRL_algorithm == 'DQN':
                model_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' +str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
                select_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'   # 選擇模型的路径            
            elif ros_topic.DRL_algorithm == 'DDQN':
                model_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
                select_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                model_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
                select_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            # 訓練
            drl.env.model_select = "train"
            drl.env.point_Workspace_cal_Monte_Carlo()
            train_env, train_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=1)
            num_iterations = 10000
            retrain_path = select_path
            retrain = ReTrainer(train_agent, train_env, model_path, num_iterations, retrain_path)
            retrain.train(train_eps = ddqn_train_eps)
            # # 測試
            drl.env.model_select = "test"
            # save_result_path = curr_path + '/test_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
            #     '/' + curr_time   # 保存結果的路径
            # plot_cfg.model_path = plot_cfg.model_path +'model_last.pkl'
            test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
            test = Tester(test_env, model_path, drl.env, num_episodes = 200) # 20230309  change 300-> 200
            test.test()
            break
        # 多種任務測試
        if ros_topic.cmd_run == 4:
            ros_topic.cmd_run = 0
            arm_structure_dof = 6
            if ros_topic.DRL_algorithm == 'DQN':
                select_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'   # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                select_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                select_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            
            drl.env.model_select = "test"
            model_path = select_path
            # test_episodes = 50

            with open(curr_path+'/muti_mission_input.yaml', 'r') as f:
                config = yaml.load(f, Loader=yaml.Loader)
            payload = config['payload']
            rospy.loginfo('Input payload: {}'.format(payload))
            work_space = config['work_space']
            rospy.loginfo('Input work_space: {}'.format(work_space))
            mission_time = config['mission_time']
            rospy.loginfo('Input mission_time: {}'.format(mission_time))
            combinations = list(itertools.product(payload, work_space, mission_time))
            rospy.loginfo('Input combinations: {}'.format(combinations))
            for p, w, t in itertools.product(payload, work_space, mission_time):
                curr_time = f'{p}-{w}-{t}'
                drl.env.op_payload = p
                # if w == 'A':  # 圓形環門
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_circle.xlsx'
                # elif w == 'B': # 複雜點位
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
                # elif w == 'C': # 複雜點位 2
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c.xlsx'
                if w == 'A':  # 圓形還門 real
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_d.xlsx'
                elif w == 'B': # 複雜點位 2 real
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real.xlsx'
                elif w == 'C': # complex real traj
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real_traj.xlsx'
                # elif w == 'D': 
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_ori_random_v2.xlsx'
                drl.env.mission_time = t
                # 指定 TensorBoard 日志的存储路径，并将作为日志文件名的一部分
                log_dir = f"logs/"+ op_function_flag +"/"+ str(ros_topic.test_model_name)+"/"+str(curr_time)+"/"
                tb = tensorboardX.SummaryWriter(log_dir = log_dir) # reset tb
                # curr_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")  # 改為任務編號
                test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
                test = Tester(test_env, model_path, drl.env, num_episodes = test_episodes) # 20230309  change 300-> 200 #20230326 mini-test 20
                test.test()
            break
        # original test
        if ros_topic.cmd_run == 5:
            with open(curr_path+'/muti_mission_input.yaml', 'r') as f:
                config = yaml.load(f, Loader=yaml.Loader)
            payload = config['payload']
            rospy.loginfo('Input payload: {}'.format(payload))
            work_space = config['work_space']
            rospy.loginfo('Input work_space: {}'.format(work_space))
            mission_time = config['mission_time']
            rospy.loginfo('Input mission_time: {}'.format(mission_time))
            combinations = list(itertools.product(payload, work_space, mission_time))
            rospy.loginfo('Input combinations: {}'.format(combinations))

            original_design = Workbook()
            sheet_original_design = original_design.active
            i = 0
            for p, w, t in itertools.product(payload, work_space, mission_time):
                curr_time = f'{p}-{w}-{t}'
                drl.env.op_payload = p
                # if w == 'A':  # 圓形環門
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_circle.xlsx'
                # elif w == 'B': # 複雜點位
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
                # elif w == 'C': 
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c.xlsx'
                if w == 'A':  # 圓形還門 real
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_d.xlsx'
                elif w == 'B': # 複雜點位 2 real
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real.xlsx'
                elif w == 'C': # complex real traj
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real_traj.xlsx'
                # elif w == 'D': 
                #     drl.env.point_test_excel = './xlsx/task_point_d.xlsx'
                drl.env.mission_time = t
                print(curr_time)
                # origin_return = drl.env.original_design(26.036,23.964,25.3,25.3,drl.env.op_payload,drl.env.mission_time)

                origin_return = drl.env.original_design(28.190,27.810,44.7,25.3,drl.env.op_payload,drl.env.mission_time)
                # origin_return = drl.env.original_design(27.56563,29.43437,25.3,25.3,drl.env.op_payload,drl.env.mission_time)
                # origin_return = drl.env.original_design(22.9566,34.0433,44.7,44.7,drl.env.op_payload,drl.env.mission_time)

                # origin_return = drl.env.original_design(26.036,23.964,44.7,44.7,drl.env.op_payload,drl.env.mission_time)
                # origin_return = drl.env.original_design(30,30,44.7,25.3,drl.env.op_payload,drl.env.mission_time)
                # 迭代矩陣的每一個元素，並填入工作表中
                for l in range(len(origin_return)):
                    sheet_original_design.cell(row=i+1, column=1).value = curr_time
                    sheet_original_design.cell(row=i+1, column=l+2).value = origin_return[l]
                i = i + 1
            file_name_original_design = "./xlsx/tested_state_original_design_" + op_function_flag + ".xlsx"
            original_design.save(file_name_original_design)
            break
        # single performance index train only for case2
        if ros_topic.cmd_run == 6:
            # drl.env.train_flag = 2
            tb = tensorboardX.SummaryWriter()
            ros_topic.cmd_run = 0
            if ros_topic.DRL_algorithm == 'DQN':
                model_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' +str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                model_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                model_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.DRL_algorithm) + '-' + str(arm_structure_dof) + '-' +str(drl.env.action_select) + '-' + curr_time + '/models/'  # 保存模型的路径
            
            # 訓練
            drl.env.model_select = "train"
            drl.env.point_Workspace_cal_Monte_Carlo()
            train_env, train_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=1)
            # model_path = None
            train = Trainer(train_agent, train_env, model_path)
            train.train(train_eps = ddqn_train_eps)

            break

        # 多種任務測試
        if ros_topic.cmd_run == 7:
            ros_topic.cmd_run = 0
            arm_structure_dof = 6
            # only manipulator, only consumption, only torque limit  
            ros_topic.test_model_name = ['DDQN-6-variable-20230529-144041','DDQN-6-variable-20230530-044116','DDQN-6-variable-20230531-054016']
            for i in range(len(ros_topic.test_model_name)):
                drl.env.train_flag = i+2
                if ros_topic.DRL_algorithm == 'DQN':
                    select_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                    '/' + str(ros_topic.test_model_name[i]) + '/models/'   # 選擇模型的路径
                elif ros_topic.DRL_algorithm == 'DDQN':
                    select_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                    '/' + str(ros_topic.test_model_name[i]) + '/models/'  # 選擇模型的路径
                elif ros_topic.DRL_algorithm == 'C51':
                    select_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                    '/' + str(ros_topic.test_model_name[i]) + '/models/'  # 選擇模型的路径
                
                drl.env.model_select = "test"
                model_path = select_path
                # test_episodes = 50

                with open(curr_path+'/muti_mission_input.yaml', 'r') as f:
                    config = yaml.load(f, Loader=yaml.Loader)
                payload = config['payload']
                rospy.loginfo('Input payload: {}'.format(payload))
                work_space = config['work_space']
                rospy.loginfo('Input work_space: {}'.format(work_space))
                mission_time = config['mission_time']
                rospy.loginfo('Input mission_time: {}'.format(mission_time))
                combinations = list(itertools.product(payload, work_space, mission_time))
                rospy.loginfo('Input combinations: {}'.format(combinations))
                for p, w, t in itertools.product(payload, work_space, mission_time):
                    curr_time = f'{p}-{w}-{t}'
                    drl.env.op_payload = p
                    # if w == 'A':  # 圓形環門
                    #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_circle.xlsx'
                    # elif w == 'B': # 複雜點位
                    #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
                    # elif w == 'C': # 複雜點位 2
                    #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c.xlsx'
                    if w == 'A':  # 圓形還門 real
                        drl.env.point_test_excel = './xlsx/task_point_6dof_tested_d.xlsx'
                    elif w == 'B': # 複雜點位 2 real
                        drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real.xlsx'
                    elif w == 'C': # complex real traj
                        drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real_traj.xlsx'
                    # elif w == 'D': 
                    #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_ori_random_v2.xlsx'
                    drl.env.mission_time = t
                    # 指定 TensorBoard 日志的存储路径，并将作为日志文件名的一部分
                    log_dir = f"logs/"+ op_function_flag +"/"+ str(ros_topic.test_model_name)+"/"+str(curr_time)+"/"
                    tb = tensorboardX.SummaryWriter(log_dir = log_dir) # reset tb
                    # curr_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")  # 改為任務編號
                    test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
                    test = Tester(test_env, model_path, drl.env, num_episodes = test_episodes) # 20230309  change 300-> 200 #20230326 mini-test 20
                    test.test()
            break
            # other robot manipulators compare
        if ros_topic.cmd_run == 8:
            with open(curr_path+'/muti_mission_input.yaml', 'r') as f:
                config = yaml.load(f, Loader=yaml.Loader)
            payload = config['payload']
            rospy.loginfo('Input payload: {}'.format(payload))
            work_space = config['work_space']
            rospy.loginfo('Input work_space: {}'.format(work_space))
            mission_time = config['mission_time']
            rospy.loginfo('Input mission_time: {}'.format(mission_time))
            combinations = list(itertools.product(payload, work_space, mission_time))
            rospy.loginfo('Input combinations: {}'.format(combinations))

            original_design = Workbook()
            sheet_original_design = original_design.active
            i = 0
            for p, w, t in itertools.product(payload, work_space, mission_time):
                curr_time = f'{p}-{w}-{t}'
                drl.env.op_payload = p
                # if w == 'A':  # 圓形環門
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_circle.xlsx'
                # elif w == 'B': # 複雜點位
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
                # elif w == 'C': 
                #     drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c.xlsx'
                if w == 'A':  # 圓形還門 real
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_d.xlsx'
                elif w == 'B': # 複雜點位 2 real
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real.xlsx'
                elif w == 'C': # complex real traj
                    drl.env.point_test_excel = './xlsx/task_point_6dof_tested_c_real_traj.xlsx'
                # elif w == 'D': 
                #     drl.env.point_test_excel = './xlsx/task_point_d.xlsx'
                drl.env.mission_time = t
                print(curr_time)
                # origin_return = drl.env.original_design(26.036,23.964,25.3,25.3,drl.env.op_payload,drl.env.mission_time)

                origin_return = drl.env.other_manipulator_design(drl.env.op_payload,drl.env.mission_time)
                # 迭代矩陣的每一個元素，並填入工作表中
                for l in range(len(origin_return)):
                    sheet_original_design.cell(row=i+1, column=1).value = curr_time
                    sheet_original_design.cell(row=i+1, column=l+2).value = origin_return[l]
                i = i + 1
            file_name_original_design = "./xlsx/denso_tested_state_original_design_" + op_function_flag + ".xlsx"
            original_design.save(file_name_original_design)
            break
        else:
            pass
